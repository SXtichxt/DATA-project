from openpyxl import load_workbook
import mysql.connector

#EXCEL
workbook = load_workbook('imported.xlsx')
sheet = workbook.active

values = []
for row in sheet.iter_rows(min_row = 2, values_only = True):
    print(row)
    values.append(row)

# DATABASE 
db = mysql.connector.connect(
    host = 'localhost',
    port = 3306,
    user = 'root',
    password = '',
    database = 'i_want_to_buy'
)

cursor = db.cursor()
sql = '''
    INSERT INTO products(title,price,necessary)
    VALUES (%s,%s,%s);
'''
cursor.executemany(sql, values)
db.commit()
print('ADD DATA '+ str(cursor.rowcount) + 'row')